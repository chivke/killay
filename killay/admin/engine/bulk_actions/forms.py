from io import BytesIO
from zipfile import BadZipFile

from django import forms
from django.core.exceptions import ValidationError
from django.urls import reverse

from openpyxl import load_workbook

from killay.admin.engine.bulk_actions.executors import PieceCreateExecutor
from killay.admin.engine.bulk_actions.serializers import PieceCreateSerializer
from killay.admin.lib.constants import BulkActionConstants


class BulkActionForm(forms.Form):
    xls_file = forms.FileField(
        label=BulkActionConstants.XLS_FILE_LABEL,
        help_text=BulkActionConstants.XLS_FILE_HELP_TEXT,
        required=True,
    )

    serializers = {BulkActionConstants.TYPE_PIECE_CREATE: PieceCreateSerializer}
    executors = {BulkActionConstants.TYPE_PIECE_CREATE: PieceCreateExecutor}

    def __init__(self, action_type: str, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.action_type = action_type

    def clean(self, *args, **kwargs) -> dict:
        cleaned_data = super().clean(*args, **kwargs)
        if "xls_file" not in cleaned_data:
            return cleaned_data
        xls_file = cleaned_data["xls_file"]
        xls_file_data = self._validate_xls_file(
            action_type=self.action_type, xls_file=xls_file
        )
        cleaned_data["xls_file_data"] = xls_file_data
        return {
            "action_type": self.action_type,
            "xls_file_data": xls_file_data,
        }

    def save(self):
        executor_class = self.executors[self.action_type]
        data_list = self.cleaned_data["xls_file_data"]
        executor = executor_class(data_list=data_list)
        return executor.run()

    def _validate_xls_file(self, action_type, xls_file) -> dict:
        data_list = self._get_file_data_list(xls_file=xls_file)
        if not data_list:
            message = BulkActionConstants.ERROR_FILE_IS_EMPTY
            raise ValidationError(
                {"xls_file": message},
            )
        assert action_type in self.serializers
        serializer_class = self.serializers[action_type]
        serializer = serializer_class(data_list=data_list)
        try:
            validated_data = serializer.validate()
        except ValidationError as error:
            row_template = BulkActionConstants.ROW_TEMPLATE
            error_list = [
                f"{row_template.format(row=row+1)}: {';'.join(messages)}"
                for row, messages in sorted(error.message_dict.items())
            ]
            raise ValidationError(error_list)
        return validated_data

    def _get_file_data_list(self, xls_file) -> list:
        try:
            workbook = load_workbook(
                filename=BytesIO(xls_file.read()),
                data_only=True,
                read_only=True,
            )
        except BadZipFile:
            raise ValidationError(
                {"xls_file": BulkActionConstants.ERROR_WRONG_FILE},
            )
        woorksheet = workbook.active
        fields = [cell.value for cell in next(woorksheet.iter_rows())]
        data_list = []
        for row in list(woorksheet.iter_rows())[1:]:
            data = {fields[index]: cell.value for index, cell in enumerate(row)}
            if any(bool(value) for field, value in data.items()):
                data_list.append(data)
        return data_list

    def get_file_headers_context(self) -> dict:
        headers_data = self.serializers[self.action_type].get_headers_data()
        required_cols = []
        non_required_cols = []
        for column_name, data in headers_data.items():
            if data["is_required"]:
                required_cols.append(data)
            else:
                non_required_cols.append(data)
        return {
            "required_cols": required_cols,
            "non_required_cols": non_required_cols,
            "title": BulkActionConstants.SECTION_COLUMNS_TITLE,
            "description": BulkActionConstants.SECTION_COLUMNS_DESCRIPTION,
            "non_required_title": BulkActionConstants.TITLE_NON_REQUIRED_COLS,
            "required_title": BulkActionConstants.TITLE_REQUIRED_COLS,
            "column_data_fields": BulkActionConstants.COLUMN_DATA_FIELDS,
            "template_title": BulkActionConstants.TITLE_TEMPLATE,
            "template_description": BulkActionConstants.DESCRIPTION_TEMPLATE,
            "template_link": reverse(
                "admin:bulk_action_template",
                kwargs={"action": self.action_type},
            ),
        }
